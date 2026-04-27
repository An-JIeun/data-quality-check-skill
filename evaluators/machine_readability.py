"""기계가독성 평가 모듈

평가 순서: (1) 기계가독형 포맷 → (2) 표준 인코딩 준수 → (3) 구조 정규성
각 단계에서 score == 0 이면 평가 실패로 간주한다.
"""

from __future__ import annotations

import os
import warnings
from typing import List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..report import EvalStatus, MetricResult

_DIMENSION = "기계가독성"


# ------------------------------------------------------------------ #
# (1) 기계가독형 포맷                                                   #
# ------------------------------------------------------------------ #
def check_format(file_path: str) -> MetricResult:
    """파일 확장자 기반 기계가독 포맷 검사.

    CSV=100 / XLSX·XLS=50 / 기타=0
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        score, status = 100, EvalStatus.PASS
        detail = "CSV 포맷 — 개방형 기계가독 포맷 (score=100)"
    elif ext in (".xlsx", ".xls"):
        score, status = 50, EvalStatus.PASS
        detail = f"{ext.upper()} 포맷 — 기계가독 가능 (score=50)"
    else:
        score, status = 0, EvalStatus.FAIL
        detail = f"평가 불가 포맷: {ext!r} (score=0)"

    return MetricResult(
        dimension=_DIMENSION,
        metric_name="기계가독형 포맷",
        scope="file",
        score=score,
        status=status,
        formula="CSV=100, XLSX/XLS=50, 기타=0",
        detail=detail,
    )


# ------------------------------------------------------------------ #
# (2) 표준 인코딩 준수                                                  #
# ------------------------------------------------------------------ #
def check_encoding(file_path: str) -> MetricResult:
    """인코딩 감지.

    UTF-8=100 / EUC-KR·CP949=50 / 기타=0
    Excel(.xlsx/.xls)은 인코딩 이슈가 없으므로 score=100 부여.
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext in (".xlsx", ".xls"):
        return MetricResult(
            dimension=_DIMENSION,
            metric_name="표준 인코딩 준수",
            scope="file",
            score=100,
            status=EvalStatus.PASS,
            formula="UTF-8=100, EUC-KR/CP949=50, 기타=0",
            detail="Excel 파일 — 인코딩 검사 해당 없음 (score=100)",
        )

    try:
        with open(file_path, "rb") as fh:
            raw = fh.read()
    except Exception as exc:  # noqa: BLE001
        return MetricResult(
            dimension=_DIMENSION,
            metric_name="표준 인코딩 준수",
            scope="file",
            score=0,
            status=EvalStatus.FAIL,
            formula="UTF-8=100, EUC-KR/CP949=50, 기타=0",
            detail=f"파일 읽기 오류: {exc} (score=0)",
        )

    _ENCODINGS_TO_TRY = [("utf-8", 100), ("cp949", 50), ("euc-kr", 50)]
    score, status, detail = 0, EvalStatus.FAIL, "인코딩 감지 실패 (score=0)"
    for enc_try, enc_score in _ENCODINGS_TO_TRY:
        try:
            raw.decode(enc_try)
            score, status = enc_score, EvalStatus.PASS
            detail = f"{enc_try} 인코딩 (score={enc_score})"
            break
        except (UnicodeDecodeError, LookupError):
            continue

    return MetricResult(
        dimension=_DIMENSION,
        metric_name="표준 인코딩 준수",
        scope="file",
        score=score,
        status=status,
        formula="UTF-8=100, EUC-KR/CP949=50, 기타=0",
        detail=detail,
    )


# ------------------------------------------------------------------ #
# (3) 구조 정규성                                                        #
# ------------------------------------------------------------------ #
def detect_column_merges(file_path: str) -> List[str]:
    """엑셀 파일에서 가로 방향 병합 셀 범위를 탐지한다.

    반환값은 ``시트명!A1:B1`` 형식의 병합 범위 목록이다.
    CSV는 병합 셀이 없으므로 빈 리스트를 반환한다.
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        return []

    merged_ranges: List[str] = []

    if ext in (".xlsx", ".xlsm"):
        try:
            with warnings.catch_warnings():
                warnings.filterwarnings(
                    "ignore",
                    message="Workbook contains no default style, apply openpyxl's default",
                    category=UserWarning,
                )
                workbook = load_workbook(file_path, read_only=False, data_only=True)
        except Exception:
            return []

        for worksheet in workbook.worksheets:
            for merged_range in worksheet.merged_cells.ranges:
                min_col, _min_row, max_col, _max_row = merged_range.bounds
                if max_col > min_col:
                    merged_ranges.append(f"{worksheet.title}!{merged_range}")
        return merged_ranges

    if ext == ".xls":
        try:
            import xlrd  # type: ignore
        except ImportError:
            return []

        try:
            workbook = xlrd.open_workbook(file_path, formatting_info=True)
        except Exception:
            return []

        for sheet in workbook.sheets():
            for row_first, row_last, col_first, col_last in sheet.merged_cells:
                if col_last - col_first > 1:
                    start = f"{get_column_letter(col_first + 1)}{row_first + 1}"
                    end = f"{get_column_letter(col_last)}{row_last}"
                    merged_ranges.append(f"{sheet.name}!{start}:{end}")

        return merged_ranges

    return []


def check_structure(
    file_path: str,
) -> Tuple[MetricResult, Optional[pd.DataFrame]]:
    """컬럼 병합 여부 및 파일 로딩 가능 여부 검사.

    병합 없음=100 / 컬럼 병합 감지=50 / 컬럼 추출 불가=0
    Returns (MetricResult, DataFrame | None)
    """
    ext = os.path.splitext(file_path)[1].lower()

    # ---- CSV --------------------------------------------------------
    if ext == ".csv":
        encodings_to_try = ["utf-8", "utf-8-sig", "euc-kr", "cp949"]
        df = None
        last_err = None
        for enc in encodings_to_try:
            try:
                df = pd.read_csv(file_path, encoding=enc)
                break
            except (UnicodeDecodeError, LookupError):
                continue
            except Exception as exc:  # noqa: BLE001
                last_err = exc
                break

        if df is not None:
            score, status = 100, EvalStatus.PASS
            detail = "컬럼 추출 가능, 행 정보 오류 없음 (score=100)"
        else:
            score, status = 0, EvalStatus.FAIL
            detail = f"파일 읽기 실패: {last_err} (score=0)"

        return (
            MetricResult(
                dimension=_DIMENSION,
                metric_name="구조 정규성",
                scope="file",
                score=score,
                status=status,
                formula="병합 없음=100, 컬럼 병합 감지=50, 컬럼 추출 불가=0",
                detail=detail,
            ),
            df,
        )

    # ---- Excel ------------------------------------------------------
    if ext in (".xlsx", ".xls"):
        df = None

        try:
            with warnings.catch_warnings():
                warnings.filterwarnings(
                    "ignore",
                    message="Workbook contains no default style, apply openpyxl's default",
                    category=UserWarning,
                )
                df = pd.read_excel(file_path)
        except Exception as exc:  # noqa: BLE001
            return (
                MetricResult(
                    dimension=_DIMENSION,
                    metric_name="구조 정규성",
                    scope="file",
                    score=0,
                    status=EvalStatus.FAIL,
                    formula="병합 없음=100, 컬럼 병합 감지=50, 컬럼 추출 불가=0",
                    detail=f"파일 읽기 실패: {exc} (score=0)",
                ),
                None,
            )

        merged_ranges = detect_column_merges(file_path)
        if merged_ranges:
            score, status = 50, EvalStatus.WARNING
            preview = ", ".join(merged_ranges[:3])
            more = f" 외 {len(merged_ranges) - 3}건" if len(merged_ranges) > 3 else ""
            detail = f"컬럼 병합 감지: {preview}{more} (score=50)"
        else:
            score, status = 100, EvalStatus.PASS
            detail = "컬럼 추출 가능, 컬럼 병합 없음 (score=100)"

        return (
            MetricResult(
                dimension=_DIMENSION,
                metric_name="구조 정규성",
                scope="file",
                score=score,
                status=status,
                formula="병합 없음=100, 컬럼 병합 감지=50, 컬럼 추출 불가=0",
                detail=detail,
            ),
            df,
        )

    # ---- 지원하지 않는 포맷 ------------------------------------------
    return (
        MetricResult(
            dimension=_DIMENSION,
            metric_name="구조 정규성",
            scope="file",
            score=0,
            status=EvalStatus.FAIL,
            formula="병합 없음=100, 컬럼 병합 감지=50, 컬럼 추출 불가=0",
            detail=f"평가 불가 포맷: {ext!r} (score=0)",
        ),
        None,
    )
